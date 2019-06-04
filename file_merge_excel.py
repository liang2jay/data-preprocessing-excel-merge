import os
import collections
import operator
from openpyxl import load_workbook
from openpyxl import Workbook


# get all of excel file from directory, DO NOT CONSIDER THE FILES SAVED IN SUBDIRECTORIES
def search_excel(source_dir, to_file):
    file_results = []
    for _root, _dirs, _files in os.walk(source_dir):
        for _file in _files:
            if _file.endswith('.xlsx'):
                file_results.append(os.path.join(_root, _file))
    try:
        print('Remove combine.xlsx.')
        file_results.remove(to_file)
    except ValueError:
        print('combine.xlsx not exist')
    return file_results


# read excel file content, return title and array
def load_excel(excel_file):
    wb = load_workbook(excel_file, read_only=True)
    ws = wb.active
    _title = []
    _items = collections.OrderedDict()
    for _row in ws.rows:
        # load and append title from first row
        if not _title:
            for _i in _row:
                _title.append(_i.value)
        else:
            # the rest of row is content
            _item = []
            for _i in _row:
                _item.append(_i.value)
            _items[_item[0]] = _item

    wb.close()
    return _title, _items


# save excel
def save_excel(excel_file, excel_title, excel_items):
    wb = Workbook()
    ws = wb.active
    ws.append(excel_title)
    for _k, _v in excel_items.items():
        ws.append(_v)
    wb.save(excel_file)


def combine(from_dir, file):
    _excel_files = search_excel(from_dir, file)
    if not _excel_files:
        return

    _excel_title = []
    _excel_content = collections.OrderedDict()
    for _file in _excel_files:
        print('Parsing ' + _file)
        _title, _items = load_excel(_file)
        if not _title or not _items:
            print('Skip since it is empty.')
            continue

        if not _excel_title:
            _excel_title = _title
        elif not operator.eq(_title, _excel_title):
            print('Warning: Excel title format are different!')

        for _k, _v in _items.items():
            _excel_content[_k] = _v
        print('Parsing done.')

    if not _excel_title or not _excel_content:
        print('All files is empty.')
    save_excel(file, _excel_title, _excel_content)


if __name__ == '__main__':
    print('begin')
    # FROM_DIR = os.getcwd()
    FROM_DIR = 'E:\github\data-preprocessing\excel\information-merge'
    TO_FILE = os.path.join(FROM_DIR, 'combine.xlsx')
    combine(FROM_DIR, TO_FILE)
    print('end')















